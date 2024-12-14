from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

import os
from django.conf import settings
from django.utils import timezone

from logging import Logger
import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from rest_framework import status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from ...optapy_solver.solver import run_optimization
from ...time_table_models import Timetable, ClassSection, Tutor, Lesson,DayChoices
from django.db.models import Prefetch
from collections import defaultdict
from rest_framework.decorators import api_view
from rest_framework.response import Response
from ...models import  Teacher, Subject, Room,Classroom,Standard
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape,letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO
import requests
from ..serializer.time_table_serializer import ClassroomWeekTimetableSerializer,TeacherWeekTimetableSerializer
import logging
from .time_table_views import get_student_condensed_timetable,get_teacher_condensed_timetable

logger = logging.getLogger(__name__)





@api_view(['GET'])
@permission_classes([IsAuthenticated])
def abbreviated_teacher_timetable_file_export(request, pk=None):
    """
    Export teacher's timetable as PDF or Excel file
    
    Query Parameters:
    - timetable_id: ID of the timetable to export (optional)
    - file_type: 'pdf' or 'xlsx'
    """
    try:
        # Extract query parameters
        file_type = request.query_params.get('file_type', 'xlsx').lower()
        
        user = request.user

        # Validate inputs
        if file_type not in ['pdf', 'xlsx']:
            return Response({
                'error': 'Invalid file type. Use "pdf" or "xlsx"'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Fetch timetable with proper access control
        try:
            if pk is None:
                # Get the default school timetable if no specific timetable is provided
                timetable = Timetable.objects.filter(school=user, is_default=True).first()
            else:
                timetable = Timetable.objects.get(school=user, id=pk)
        except Timetable.DoesNotExist:
            return Response({
                'error': 'Timetable not found or access denied'
            }, status=status.HTTP_404_NOT_FOUND)

        # Generate condensed timetable data for the teacher
        condensed_timetable = get_teacher_condensed_timetable( timetable)
        
        # Prepare weekly schedule header
        weekly_schedule_header = [
            {
                "day": schedule.day,
                "teaching_slots": schedule.teaching_slots,
                "day_name": dict(DayChoices.choices).get(schedule.day)
            }
            for schedule in timetable.day_schedules.all()
        ]

        # Prepare filename with timestamp
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"teacher_timetable_{user.school_name}_{timestamp}.{file_type}"

        # Ensure media directory exists
        export_dir = os.path.join(settings.MEDIA_ROOT, 'timetables')
        os.makedirs(export_dir, exist_ok=True)

        # Full file path
        file_path = os.path.join(export_dir, filename)

        # Generate file based on type
        try:
            if file_type == 'pdf':
                filename = f"{user.school_name}_teacher_whole_week_timetable.pdf"
                response = generate_whole_teacher_full_week_timetable_pdf(
                    timetable, 
                    request.user, 
                    condensed_timetable, 
                    weekly_schedule_header, 
                    file_path
                )
            else:
                filename = f"{user.school_name}_teacher_whole_week_timetable.xlsx"
                response = generate_whole_teacher_full_week_timetable_excel(
                    timetable, 
                    request.user, 
                    condensed_timetable, 
                    weekly_schedule_header, 
                    file_path
                )

            return response
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    except Exception as e:
        # Log the error for debugging
        logger.error(f"Teacher timetable export error: {str(e)}")
        import traceback
        error_details = traceback.format_exc()  # Captures the full traceback
        print("error_details", error_details)
        return Response({
            'error': 'An unexpected error occurred during export',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def generate_whole_teacher_full_week_timetable_pdf(timetable, user, condensed_timetable, weekly_schedule_header, file_path):
    """
    Generate a visually appealing PDF timetable for teachers
    
    :param timetable: Timetable instance
    :param user: User downloading the timetable
    :param condensed_timetable: Processed timetable data
    :param weekly_schedule_header: Day schedule information
    :param file_path: Path to save the PDF
    :return: HttpResponse with PDF
    """
    
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    # Create buffer for PDF
    buffer = BytesIO()

    # PDF generation with custom page setup
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                            leftMargin=0.5*inch, 
                            rightMargin=0.5*inch, 
                            topMargin=0.5*inch, 
                            bottomMargin=0.5*inch)
    
    # Get sample styles
    styles = getSampleStyleSheet()
    
    # Modify title style
    title_style = styles['Title'].clone('CustomTitle')
    title_style.textColor = colors.navy
    title_style.fontSize = 16
    title_style.alignment = 1  # Center alignment

    # Story to build PDF
    story = []

    # Title (safely handle None)
    username = user.full_name if hasattr(user, 'full_name') else user.username
    title = Paragraph(f"Teacher Timetable - {username}", title_style)
    story.append(title)
    story.append(Spacer(1, 12))

    # Prepare table header
    header_row = ['Periods'] + [day['day_name'] for day in weekly_schedule_header]
    table_data = [header_row]

    # Prepare the timetable data
    for period_index in range(max(day['teaching_slots'] for day in weekly_schedule_header)):
        row_data = [f"Period {period_index + 1}"]
        
        # Process each day
        for day_schedule in weekly_schedule_header:
            # Check if this period exists for the day
            if period_index < day_schedule['teaching_slots']:
                day_sessions = condensed_timetable['timetable_rows'].get(day_schedule['day'], [])
                
                # Get session for this specific period
                period_sessions = day_sessions[period_index] if period_index < len(day_sessions) else []
                
                # Combine session details
                if period_sessions:
                    session_details = [
                        f"{session['subject']} (Room: {session['room_no']})" 
                        for session in period_sessions
                    ]
                    row_data.append('\n'.join(session_details))
                else:
                    row_data.append('-')
            else:
                row_data.append('-')
        
        table_data.append(row_data)

    # Create table with enhanced styling
    table = Table(table_data, repeatRows=1, colWidths=[1.5*inch] + [1.5*inch]*len(weekly_schedule_header))
    table.setStyle(TableStyle([
        # Header row styling
        ('BACKGROUND', (0,0), (-1,0), colors.navy),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        
        # Period column styling
        ('BACKGROUND', (0,1), (0,-1), colors.lightblue),
        ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'),
        
        # Body cells styling
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        
        # Grid styling
        ('GRID', (0,0), (-1,-1), 1, colors.lightgrey),
        ('BOX', (0,0), (-1,-1), 2, colors.navy),
    ]))

    story.append(table)

    # Build PDF
    doc.build(story)
    
    # Seek to the beginning of the buffer
    buffer.seek(0)
    
    # Prepare response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="teacher_timetable.pdf"'
    
    return response


def generate_whole_teacher_full_week_timetable_excel(timetable, user, condensed_timetable, weekly_schedule_header, file_path):
    """
    Generate Excel timetable for teachers with improved styling
    
    :param timetable: Timetable instance
    :param user: User downloading the timetable
    :param condensed_timetable: Processed timetable data
    :param weekly_schedule_header: Day schedule information
    :param file_path: Path to save the Excel file
    :return: FileResponse
    """
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teacher Timetable"

    # Color definitions to match frontend
    BLUE_HEADER_BG = '2563EB'  # Tailwind blue-600
    WHITE_TEXT = 'FFFFFF'
    BLUE_LIGHT_BG = 'DBEAFE'  # Tailwind blue-100
    BLUE_TEXT = '1E40AF'      # Tailwind blue-800
    PURPLE_ELECTIVE_BG = 'E9D5FF'  # Tailwind purple-200
    PURPLE_ELECTIVE_TEXT = '7E22CE'  # Tailwind purple-900
    BLUE_SESSION_BG = 'BFDBFE'  # Tailwind blue-200
    BLUE_SESSION_TEXT = '1E40AF'  # Tailwind blue-900

    # First row: Main header with days
    header_row = ['Periods'] + [day['day_name'] for day in weekly_schedule_header]
    ws.append(header_row)
    
    # Style first row (main header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color=WHITE_TEXT)
        cell.fill = PatternFill(start_color=BLUE_HEADER_BG, end_color=BLUE_HEADER_BG, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Prepare the timetable data
    max_periods = max(day['teaching_slots'] for day in weekly_schedule_header)
    
    for period_index in range(max_periods):
        row_data = [f"Period {period_index + 1}"]
        
        # Process each day
        for day_schedule in weekly_schedule_header:
            # Check if this period exists for the day
            if period_index < day_schedule['teaching_slots']:
                day_sessions = condensed_timetable['timetable_rows'].get(day_schedule['day'], [])
                
                # Get session for this specific period
                period_sessions = day_sessions[period_index] if period_index < len(day_sessions) else []
                
                # Combine session details
                if period_sessions:
                    session_texts = []
                    for session in period_sessions:
                        session_type = 'Elective' if session.get('is_elective') else 'Mandatory'
                        session_texts.append(f"{session['subject']} (Room: {session['room_no']}, {session_type})")
                    row_data.append('\n'.join(session_texts))
                else:
                    row_data.append('')
            else:
                row_data.append('')
        
        ws.append(row_data)

    # Styling for data rows
    for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Style period identifier column
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill(start_color=BLUE_LIGHT_BG, end_color=BLUE_LIGHT_BG, fill_type='solid')

        # Style data cells
        for cell_index, cell in enumerate(row[1:], start=1):
            # Determine if it's an elective or mandatory session
            if cell.value:
                is_elective = 'Elective' in str(cell.value)
                
                if is_elective:
                    cell.fill = PatternFill(start_color=PURPLE_ELECTIVE_BG, end_color=PURPLE_ELECTIVE_BG, fill_type='solid')
                    cell.font = Font(color=PURPLE_ELECTIVE_TEXT)
                else:
                    cell.fill = PatternFill(start_color=BLUE_SESSION_BG, end_color=BLUE_SESSION_BG, fill_type='solid')
                    cell.font = Font(color=BLUE_SESSION_TEXT)

            # Consistent cell styling
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 3)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(buffer) 
    buffer.seek(0)
    from django.http import FileResponse

    return FileResponse(
        buffer, 
        as_attachment=True, 
        filename=f"{user.username}_teacher_timetable.xlsx",
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )








@api_view(['GET'])
@permission_classes([IsAuthenticated])
def abbreviated_student_timetable_file_export(request,pk=None):
    """
    Export timetable as PDF or Excel file
    
    Query Parameters:
    - timetable_id: ID of the timetable to export
    - file_type: 'pdf' or 'xlsx'
    """
    try:
        # Extract query parameters
        file_type = request.query_params.get('file_type', 'xlsx').lower()
        print(file_type)
        
        user = request.user

        # Validate inputs
        

        if file_type not in ['pdf', 'xlsx']:
            return Response({
                'error': 'Invalid file type. Use "pdf" or "xlsx"'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Fetch timetable with proper access control
        try:
            if pk is None:
                 timetable = Timetable.objects.filter(school=user, is_default=True).first()  # Get the default timetable or None
            else:
                 timetable = Timetable.objects.get(school=user, id=pk)
        except Timetable.DoesNotExist:
            return Response({
                'error': 'Timetable not found or access denied'
            }, status=status.HTTP_404_NOT_FOUND)

        # Generate condensed timetable data
        condensed_timetable = get_student_condensed_timetable(request.user, timetable)
        weekly_schedule_header = [
            {
                "day": schedule.day,
                "teaching_slots": schedule.teaching_slots,
                "day_name": dict(DayChoices.choices).get(schedule.day)
            }
            for schedule in timetable.day_schedules.all()
        ]

        # Prepare filename with timestamp
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"timetable_{request.user.username}_{timestamp}.{file_type}"

        # Ensure media directory exists
        export_dir = os.path.join(settings.MEDIA_ROOT, 'timetables')
        os.makedirs(export_dir, exist_ok=True)

        # Full file path
        file_path = os.path.join(export_dir, filename)

        # Generate file based on type
        try:
            if file_type == 'pdf':
                filename = f"{user.school_name}student_whole_week_timetable.pdf"
                response =    generate_whole_student_full_week_timetable_pdf(
                timetable, 
                request.user, 
                condensed_timetable, 
                weekly_schedule_header, 
                file_path
            )
            else:
                filename = f"{user.school_name}student_whole_week_timetable.xlsx"
                response = generate_whole_student_full_week_timetable_excel(
                timetable, 
                request.user, 
                condensed_timetable, 
                weekly_schedule_header, 
                file_path
            )

            return response
        except Exception as e:
          return Response({'error': str(e)}, status=500)

    except Exception as e:
        # Log the error for debugging
        logger.error(f"Timetable export error: {str(e)}")
        
        return Response({
            'error': 'An unexpected error occurred during export',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def generate_whole_student_full_week_timetable_pdf(timetable, user, condensed_timetable, weekly_schedule_header, file_path):
    """
    Generate a visually appealing PDF timetable
    
    :param timetable: Timetable instance
    :param user: User downloading the timetable
    :param condensed_timetable: Processed timetable data
    :param weekly_schedule_header: Day schedule information
    :param file_path: Path to save the PDF
    :return: HttpResponse with PDF
    """
    
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    # Create buffer for PDF
    buffer = BytesIO()

    # PDF generation with custom page setup
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                            leftMargin=0.5*inch, 
                            rightMargin=0.5*inch, 
                            topMargin=0.5*inch, 
                            bottomMargin=0.5*inch)
    
    # Get sample styles
    styles = getSampleStyleSheet()
    
    # Modify title style
    title_style = styles['Title'].clone('CustomTitle')
    title_style.textColor = colors.navy
    title_style.fontSize = 16
    title_style.alignment = 1  # Center alignment

    # Story to build PDF
    story = []

    # Title (safely handle None)
    username = getattr(user, 'username', 'Student')
    title = Paragraph(f"Student Timetable - {username}", title_style)
    story.append(title)
    story.append(Spacer(1, 12))

    # Prepare table header
    header_row = ['Classrooms'] + [day['day_name'] for day in weekly_schedule_header]
    table_data = [header_row]

    # Process each classroom's timetable
    for classroom_data in condensed_timetable:
        row_data = [classroom_data['class_details']['full_identifier']]
        
        # Process each day
        for day_schedule in weekly_schedule_header:
            day_sessions = classroom_data['timetable_rows'].get(day_schedule['day'], [])
            
            # Combine all sessions for the day
            day_subjects = []
            for period_sessions in day_sessions:
                period_subjects = [
                    session['subject'] for session in period_sessions
                ]
                day_subjects.append(', '.join(period_subjects) if period_subjects else '-')
            
            row_data.append('\n'.join(day_subjects))
        
        table_data.append(row_data)

    # Create table with enhanced styling
    table = Table(table_data, repeatRows=1, colWidths=[1.5*inch] + [1.5*inch]*len(weekly_schedule_header))
    table.setStyle(TableStyle([
        # Header row styling
        ('BACKGROUND', (0,0), (-1,0), colors.navy),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        
        # Classroom column styling
        ('BACKGROUND', (0,1), (0,-1), colors.lightblue),
        ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'),
        
        # Body cells styling
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        
        # Grid styling
        ('GRID', (0,0), (-1,-1), 1, colors.lightgrey),
        ('BOX', (0,0), (-1,-1), 2, colors.navy),
    ]))

    story.append(table)

    # Build PDF
    doc.build(story)
    
    # Seek to the beginning of the buffer
    buffer.seek(0)
    
    # Prepare response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="student_timetable.pdf"'
    
    return response


  
def generate_whole_student_full_week_timetable_excel(timetable, user, condensed_timetable, weekly_schedule_header, file_path):
    """
    Generate Excel timetable with improved styling to match frontend design
    
    :param timetable: Timetable instance
    :param user: User downloading the timetable
    :param condensed_timetable: Processed timetable data
    :param weekly_schedule_header: Day schedule information
    :param file_path: Path to save the Excel file
    :return: FileResponse
    """
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Timetable"

    # Color definitions to match frontend
    BLUE_HEADER_BG = '2563EB'  # Tailwind blue-600
    WHITE_TEXT = 'FFFFFF'
    BLUE_LIGHT_BG = 'DBEAFE'  # Tailwind blue-100
    BLUE_TEXT = '1E40AF'      # Tailwind blue-800
    PURPLE_ELECTIVE_BG = 'E9D5FF'  # Tailwind purple-200
    PURPLE_ELECTIVE_TEXT = '7E22CE'  # Tailwind purple-900
    BLUE_SESSION_BG = 'BFDBFE'  # Tailwind blue-200
    BLUE_SESSION_TEXT = '1E40AF'  # Tailwind blue-900

    # First row: Main header with classrooms and days
    header_row = ['Classrooms'] + [day['day_name'] for day in weekly_schedule_header]
    ws.append(header_row)
    
    # Style first row (main header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color=WHITE_TEXT)
        cell.fill = PatternFill(start_color=BLUE_HEADER_BG, end_color=BLUE_HEADER_BG, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Second row: Period numbers
    period_row = ['']  # Empty cell for classroom column
    for day_schedule in weekly_schedule_header:
        period_row.extend([str(i+1) for i in range(day_schedule['teaching_slots'])])
    ws.append(period_row)

    # Style second row (period numbers)
    for cell in ws[2]:
        cell.font = Font(color=BLUE_TEXT)
        cell.fill = PatternFill(start_color=BLUE_LIGHT_BG, end_color=BLUE_LIGHT_BG, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Process each classroom's timetable
    for classroom_data in condensed_timetable:
        row_data = [classroom_data['class_details']['full_identifier']]
        
        for day_schedule in weekly_schedule_header:
            day_sessions = classroom_data['timetable_rows'].get(day_schedule['day'], [])
            
            # Ensure we have enough slots filled
            while len(day_sessions) < day_schedule['teaching_slots']:
                day_sessions.append([])
            
            for period_sessions in day_sessions:
                if not period_sessions:
                    row_data.append('')
                else:
                    # Combine sessions for the period
                    session_texts = []
                    for session in period_sessions:
                        session_type = 'Elective' if session.get('is_elective') else 'Mandatory'
                        session_texts.append(f"{session['subject']} ({session_type})")
                    row_data.append('\n'.join(session_texts))
        
        ws.append(row_data)

    # Styling for data rows
    for row_index, row in enumerate(ws.iter_rows(min_row=3), start=3):
        # Style classroom identifier column
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill(start_color=BLUE_LIGHT_BG, end_color=BLUE_LIGHT_BG, fill_type='solid')

        # Style data cells
        for cell_index, cell in enumerate(row[1:], start=1):
            # Determine if it's an elective or mandatory session
            if cell.value:
                is_elective = 'Elective' in str(cell.value)
                
                if is_elective:
                    cell.fill = PatternFill(start_color=PURPLE_ELECTIVE_BG, end_color=PURPLE_ELECTIVE_BG, fill_type='solid')
                    cell.font = Font(color=PURPLE_ELECTIVE_TEXT)
                else:
                    cell.fill = PatternFill(start_color=BLUE_SESSION_BG, end_color=BLUE_SESSION_BG, fill_type='solid')
                    cell.font = Font(color=BLUE_SESSION_TEXT)

            # Consistent cell styling
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 3)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(buffer) 
    buffer.seek(0)
    from django.http import FileResponse

    return FileResponse(
        buffer, 
        as_attachment=True, 
        filename=f"{user.username}_timetable.xlsx",
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )






@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_classroom_timetable(request, pk):
    """
    Download classroom timetable in either Excel or PDF format.
    Use ?format=pdf query parameter for PDF format.
    """
    user = request.user
    timetable = Timetable.objects.filter(school=user, is_default=True).first()  # Get the default timetable or None

    if not timetable:
        return Response({"warning": "You have no default timetable.", "data": []}, status=200)  # Return warning with empty data
    file_type = request.query_params.get('file_type', '').lower()

    try:
        classroom = get_object_or_404(Classroom, id=pk, school=user)
        classsection = get_object_or_404(ClassSection, classroom=classroom, timetable=timetable)
    except Classroom.DoesNotExist:
        return Response({'error': 'No classroom found for this school.'}, status=404)
    except ClassSection.DoesNotExist:
        return Response({'error': 'No classroom found for this timetable.'}, status=422)

    day_schedules = timetable.day_schedules.all() if timetable else []
    day_timetable = []

    for day_schedule in day_schedules:
        day_lessons = Lesson.objects.filter(
            timetable=timetable,
            timeslot__day_of_week=day_schedule.day,
            class_sections__in=[classsection]
        ).select_related(
            'course', 'allotted_teacher__teacher', 'classroom_assignment__room', 'timeslot'
        )
        
        sessions = defaultdict(list)
        for lesson in day_lessons:
            sessions[lesson.timeslot.period - 1].append(lesson)

        formatted_sessions = [sessions[i] for i in range(day_schedule.teaching_slots)]

        day_timetable.append({
            'day': day_schedule.day,
            'sessions': formatted_sessions
        })

    serializer = ClassroomWeekTimetableSerializer(day_timetable, many=True, context={'class_section': classsection})
    timetable_data = serializer.data

    try:
        if file_type == 'pdf':
            filename = f"{classroom.standard.short_name}-{classroom.division}_timetable.pdf"
            response = generate_classroom_pdf_timetable(timetable_data, classroom, filename)
        else:
            filename = f"{classroom.standard.short_name}-{classroom.division}_timetable.xlsx"
            response = generate_classroom_excel_timetable(timetable_data, classroom, filename)
        
        return response
    except Exception as e:
        return Response({'error': str(e)}, status=500)

def generate_classroom_pdf_timetable(timetable_data, classroom, filename):
    # Create PDF buffer
    buffer = BytesIO()
    
    # Create the PDF document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,
        spaceAfter=30
    )
    cell_style = ParagraphStyle(
        'CustomCell',
        parent=styles['Normal'],
        fontSize=9,
        alignment=1,
        leading=12
    )

    elements = []

    # Add title
    title = Paragraph(
        f"Class Timetable: {classroom.standard.short_name}-{classroom.division}",
        title_style
    )
    elements.append(title)
    elements.append(Spacer(1, 20))

    # Day names mapping
    day_names = {
        'MON': 'Monday',
        'TUE': 'Tuesday',
        'WED': 'Wednesday',
        'THU': 'Thursday',
        'FRI': 'Friday',
        'SAT': 'Saturday',
        'SUN': 'Sunday'
    }

    # Prepare table data
    max_sessions = max(len(day['sessions']) for day in timetable_data)

    headers = ['Day'] + [f'Period {i+1}' for i in range(max_sessions)]
    table_data = [headers]

    for day_data in timetable_data:
        row = [day_names.get(day_data['day'], day_data['day'])]
        
        for session in day_data['sessions']:
            if session:
                cell_text = []
                for distribution in session['class_distribution']:
                    subject = session['name']
                    teacher = distribution['teacher']['name']
                    room = f"{distribution['room']['name']} ({distribution['room']['room_number']})"
                    
                    if session['type'] == 'Elective':
                        students = distribution['number_of_students_from_this_class']
                        cell_text.append(f"📚 {subject}\n👨‍🏫 {teacher}\n🏫 {room}\n👥 Students: {students}")
                    else:
                        cell_text.append(f"📚 {subject}\n👨‍🏫 {teacher}\n🏫 {room}")
                
                row.append(Paragraph('\n'.join(cell_text), cell_style))
            else:
                row.append(Paragraph('-', cell_style))
        
        table_data.append(row)

    # Create table with specific column widths
    col_widths = [1.2*inch] + [1.5*inch] * (len(headers)-1)
    table = Table(table_data, colWidths=col_widths)
    
    # Define table style
    style = TableStyle([
        # Headers
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        
        # Day column
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#B4C6E7')),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1F4E79')),
        
        # Alignment
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Row heights
        ('ROWHEIGHT', (0, 0), (-1, 0), 30),
        ('ROWHEIGHT', (0, 1), (-1, -1), 80),
    ])
    
    # Add alternating row colors
    for i in range(1, len(table_data), 2):
        style.add('BACKGROUND', (1, i), (-1, i), colors.HexColor('#F5F5F5'))
    
    table.setStyle(style)
    elements.append(table)

    # Build PDF document
    doc.build(elements)
    
    # Prepare response
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def generate_classroom_excel_timetable(timetable_data, classroom, filename):
    # Your existing Excel generation code here...
    # [Keep your current Excel generation code as it works well]
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{classroom.standard.short_name}-{classroom.division} Timetable"

    # Define styles
    header_fill = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    elective_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # light yellow
    core_fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # light cyan

    # Write headers
    headers = ['Day'] + [f'Session {i+1}' for i in range(len(timetable_data[0]['sessions']))]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Write timetable data
    for row, day_data in enumerate(timetable_data, start=2):
        day_cell = ws.cell(row=row, column=1, value=day_data['day'])
        day_cell.border = border
        day_cell.font = Font(bold=True, size=12)
        day_cell.alignment = Alignment(horizontal='center', vertical='center')

        for col, session in enumerate(day_data['sessions'], start=2):
            if session:
                subject = session['name']
                session_type = session['type']
                for distribution in session['class_distribution']:
                    teacher = distribution['teacher']['name']
                    room = f"{distribution['room']['name']} ({distribution['room']['room_number']})"

                    # Style elective vs. core subjects differently
                    if session_type == 'Elective':
                        students = distribution['number_of_students_from_this_class']
                        session_info = f"Subject: {subject} (Elective)\nTeacher: {teacher}\nRoom: {room}\nStudents: {students}"
                        cell_fill = elective_fill
                    else:
                        session_info = f"Subject: {subject} (Core)\nTeacher: {teacher}\nRoom: {room}"
                        cell_fill = core_fill

                    # Create and style the cell for this session
                    cell = ws.cell(row=row, column=col, value=f"\n{session_info}\n")
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell.border = border
                    cell.fill = cell_fill

                    # Apply different styles to different parts of the content
                    lines = cell.value.split('\n')
                    for i, line in enumerate(lines):
                        if line.startswith('Subject:'):
                            lines[i] = f"$BOLD_BLUE${line}$ENDBOLD_BLUE$"
                        elif line.startswith('Teacher:'):
                            lines[i] = f"$BOLD_GREEN${line}$ENDBOLD_GREEN$"
                        elif line.startswith('Room:'):
                            lines[i] = f"$RED${line}$ENDRED$"
                        elif line.startswith('Students:'):
                            lines[i] = f"$INDIGO${line}$ENDINDIGO$"
                    
                    cell.value = "\n".join(lines)

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30  # Increased width for better spacing

    # Adjust row height to fit content
    for row in range(1, len(timetable_data) + 2):
        ws.row_dimensions[row].height = 100  # Increased height for better spacing

    # Apply rich text formatting
    for row in ws.iter_rows(min_row=2, max_row=len(timetable_data) + 1, min_col=2):
        for cell in row:
            if cell.value:
                parts = []
                current_text = ""
                current_format = None
                for part in cell.value.split('$'):
                    if part == 'BOLD_BLUE':
                        if current_text:
                            parts.append((current_text, current_format))
                        current_format = Font(bold=True, size=12, color="000080")
                        current_text = ""
                    elif part == 'ENDBOLD_BLUE':
                        parts.append((current_text, current_format))
                        current_format = None
                        current_text = ""
                    elif part == 'BOLD_GREEN':
                        if current_text:
                            parts.append((current_text, current_format))
                        current_format = Font(bold=True, size=11, color="006400")
                        current_text = ""
                    elif part == 'ENDBOLD_GREEN':
                        parts.append((current_text, current_format))
                        current_format = None
                        current_text = ""
                    elif part == 'RED':
                        if current_text:
                            parts.append((current_text, current_format))
                        current_format = Font(size=11, color="8B0000")
                        current_text = ""
                    elif part == 'ENDRED':
                        parts.append((current_text, current_format))
                        current_format = None
                        current_text = ""
                    elif part == 'INDIGO':
                        if current_text:
                            parts.append((current_text, current_format))
                        current_format = Font(size=11, color="4B0082")
                        current_text = ""
                    elif part == 'ENDINDIGO':
                        parts.append((current_text, current_format))
                        current_format = None
                        current_text = ""
                    else:
                        current_text += part
                if current_text:
                    parts.append((current_text, current_format))
                
                cell.value = parts[0][0]
                cell.font = parts[0][1] or Font(size=11)
                for text, font in parts[1:]:
                    if font:
                        cell.font = font
                    cell.value += text
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response













@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_teacher_timetable(request, pk=None):
    user = request.user
    timetable = Timetable.objects.filter(school=user, is_default=True).first()  # Get the default timetable or None

    if not timetable:
        return Response({"warning": "You have no default timetable.", "data": []}, status=200)  # Return warning with empty data

    if pk is not None:
        try:
            teacher = Teacher.objects.get(id=pk, school=user)
            tutor = Tutor.objects.get(teacher=teacher, timetable=timetable)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found for this school.'}, status=status.HTTP_404_NOT_FOUND)
        except Tutor.DoesNotExist:
            return Response({'error': 'No timetable available for this teacher.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

    day_schedules = timetable.day_schedules.all() if timetable else []
    day_timetable = []

    for day_schedule in day_schedules:
        sessions = [None] * day_schedule.teaching_slots
        lessons = Lesson.objects.filter(
            timetable=timetable,
            allotted_teacher=tutor,
            timeslot__day_of_week=day_schedule.day
        ).order_by('timeslot__period')

        for lesson in lessons:
            sessions[lesson.timeslot.period - 1] = lesson

        if sessions:
            day_timetable.append({
                'day': day_schedule.day,
                'sessions': sessions
            })

    serializer = TeacherWeekTimetableSerializer(day_timetable, many=True)
    timetable_data = serializer.data

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{teacher.name}'s Timetable"

    # Define styles
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    day_font = Font(name='Arial', bold=True, size=11)
    content_font = Font(name='Arial', size=10)
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Color mapping for session types
    color_map = {
        'Core': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # Light green
        'Elective': PatternFill(start_color="E2EFFF", end_color="E2EFFF", fill_type="solid"),  # Light blue
        'Free': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
    }
    day_fill = PatternFill(start_color="008080", end_color="66CDAA", fill_type="solid")  # Teal to light teal

    # Define the alignment (centered both vertically and horizontally)
    day_alignment = Alignment(horizontal='center', vertical='center')
    # Write headers
    headers = ["Day"] + [f"Session {i+1}" for i in range(day_schedule.teaching_slots)]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = border

    # Write timetable data
    for row, day_data in enumerate(timetable_data, start=2):
        cell = ws.cell(row=row, column=1, value=day_data['day'])
        cell.font = day_font
        cell.fill = day_fill  # Set the background color
        cell.alignment = day_alignment
            
        for col, session in enumerate(day_data['sessions'], start=2):
            cell = ws.cell(row=row, column=col)
            if session['subject']:
                cell_value = f"{session['subject'] or session['elective_subject_name']}\n"
                cell_value += f"Room: {session['room']['room_number'] if session['room'] else 'N/A'}\n"
                if session['class_details']: 
                    for class_detail in session['class_details']:
                        cell_value += f"{class_detail['standard']} {class_detail['division']}"
                        if session['type'] == 'Elective':
                            cell_value += f" ({class_detail['number_of_students']} students)"
                        cell_value += "\n"
                cell.value = cell_value.strip()
                if session['type']: 

                    cell.fill = color_map[session['type']]
            else:
                cell.value = "Free Period/\nPlanning Period"
                cell.fill = color_map['Free']
            
            cell.font = content_font
            cell.alignment = centered_alignment
            cell.border = border

    # Adjust column widths and row heights
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    for row in range(1, len(timetable_data) + 2):  # +2 to include header row
        ws.row_dimensions[row].height = 80  # Adjust this value as needed

    # Create the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{teacher.name}_timetable.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response





