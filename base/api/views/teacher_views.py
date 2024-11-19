from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from ...models import Teacher, Subject, Grade
from ..serializer.teacher_serializer import TeacherSerializer
from django.db.models import Count
from django.shortcuts import get_object_or_404
# views.py
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import base64
from django.core.files.base import ContentFile
from io import BytesIO
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import load_workbook
from django.db import transaction
import pandas as pd
from fuzzywuzzy import fuzz  # For fuzzy string matching of subject names



@api_view(['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def teacher(request,pk=None):
    if request.method == 'POST':
        
        serializer = TeacherSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            teacher = serializer.save(school=request.user)
            if 'profile_image' in request.FILES:
                teacher.profile_image = request.FILES['profile_image']
                teacher.save()
            return Response(TeacherSerializer(teacher).data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        teachers = Teacher.objects.filter(school=request.user)
        serializer = TeacherSerializer(teachers, many=True)
        return Response(serializer.data)

    elif request.method == 'PUT' and pk is not None:
        try:
            teacher = Teacher.objects.get(id=pk, school=request.user)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)
        print(request.data)

        serializer = TeacherSerializer(teacher, data=request.data, partial=True,context={'request': request})
        if serializer.is_valid():
            teacher = serializer.save()
            if 'profile_image' in request.FILES:
                teacher.profile_image = request.FILES['profile_image']
                teacher.save()
            return Response(TeacherSerializer(teacher).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        try:
            teacher = Teacher.objects.get(id=pk, school=request.user)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

        teacher.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_teacher_image(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    
    # Check if the authenticated user has permission to update this teacher
    # You might want to add more specific permission checks here
    
    if 'profile_image' not in request.FILES:
        return Response({'error': 'No image file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    teacher.profile_image = request.FILES['profile_image']
    teacher.save()
    
    serializer = TeacherSerializer(teacher)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def subject_teacher_count(request):
    school=request.user
    subject_data=school.subjects.annotate(count=Count('qualified_teachers')).values('name','count')
    formated_data=[
        {
            "subject":item["name"],
            "count":item["count"]
        }
            for item in subject_data
 
    ]
    
    return Response({'subjectData': formated_data})



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def teachers(request):
    try:
        user = request.user
        teachers = Teacher.objects.filter(school=user)
        serializer = TeacherSerializer(teachers, many=True)
        
        if not teachers.exists():
            return Response([], status=status.HTTP_200_OK)

        return Response(serializer.data, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)    
    
    
    



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_teacher_template(request):
    """Generate Excel template with working multi-select dropdowns"""
    
    # Get user-specific data
    user_grades = Grade.objects.filter(school=request.user).values_list('name', flat=True)
    user_subjects = Subject.objects.filter(school=request.user).values_list('name', flat=True)
    
    if not user_grades:
        return Response({"error": "Please create at least one grade before generating template."}, 
                      status=400)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"
    
    # Create validation sheet for dropdowns
    validation_sheet = wb.create_sheet(title="Dropdowns")
    validation_sheet.sheet_state = 'hidden'  # Hide the validation sheet
    
    # Prepare grades list with concatenated options
    # This creates a list of all possible combinations for multi-select
    grade_list = list(user_grades)
    combined_grades = []
    
    # Create all possible combinations of grades (up to 3 for performance)
    for i in range(len(grade_list)):
        combined_grades.append(grade_list[i])
        for j in range(i + 1, len(grade_list)):
            combined_grades.append(f"{grade_list[i]}, {grade_list[j]}")
            for k in range(j + 1, len(grade_list)):
                combined_grades.append(f"{grade_list[i]}, {grade_list[j]}, {grade_list[k]}")
    
    # Write grades to validation sheet
    validation_sheet['A1'] = "Grade Combinations"
    for i, grade in enumerate(combined_grades, 2):
        validation_sheet[f'A{i}'] = grade
    
    # Write subjects to validation sheet
    validation_sheet['B1'] = "Subjects"
    for i, subject in enumerate(user_subjects, 2):
        validation_sheet[f'B{i}'] = subject
    
    # Style definitions
    header_style = {
        'fill': PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        'font': Font(color="FFFFFF", bold=True, size=12),
        'border': Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        ),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
    
    # Define columns
    columns = [
        {'name': 'name*', 'width': 20, 'comment': 'Teacher\'s first name (Required)'},
        {'name': 'surname', 'width': 20, 'comment': 'Teacher\'s last name'},
        {'name': 'email', 'width': 30, 'comment': 'Valid email address'},
        {'name': 'phone', 'width': 15, 'comment': 'Contact number'},
        {'name': 'qualified_subjects*', 'width': 40, 
         'comment': 'Select from dropdown or type new subjects\nSeparate multiple with commas'},
        {'name': 'grades*', 'width': 40, 
         'comment': 'Select from dropdown (includes multiple combinations)\nOr type grades separated by commas'},
        {'name': 'min_lessons_per_week*', 'width': 20, 
         'comment': 'Minimum lessons per week (Required)'},
        {'name': 'max_lessons_per_week*', 'width': 20, 
         'comment': 'Maximum lessons per week (Required)'},
        {'name': 'teacher_id', 'width': 15, 
         'comment': 'Custom ID (optional)'},
    ]
    
    # Set up headers
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column['name'])
        ws.column_dimensions[get_column_letter(col_idx)].width = column['width']
        
        for style_attr, style_value in header_style.items():
            setattr(cell, style_attr, style_value)
        
        cell.comment = Comment(column['comment'], 'Template Guide')
    
    # Add data validations
    # Grades validation with multi-select combinations
    grades_validation = DataValidation(
        type="list",
        formula1=f"Dropdowns!$A$2:$A${len(combined_grades)+1}",
        allow_blank=False,
        showDropDown=True,
        errorTitle='Invalid Grade Selection',
        error='Please select grades from the dropdown list or type valid grades separated by commas.',
        promptTitle='Grade Selection',
        prompt='Select grade combination or type grades separated by commas'
    )
    ws.add_data_validation(grades_validation)
    grades_validation.add('F2:F1000')
    
    # Subjects validation (allowing new entries)
    subjects_validation = DataValidation(
        type="list",
        formula1=f"Dropdowns!$B$2:$B${len(user_subjects)+1}",
        allow_blank=False,
        showDropDown=True,
        errorTitle='Subject Selection',
        error='Select existing subjects or type new ones (separate with commas)',
        promptTitle='Subject Selection',
        prompt='Select or type subjects (separate multiple with commas)'
    )
    ws.add_data_validation(subjects_validation)
    subjects_validation.add('E2:E1000')
    
    # Numeric validation for lessons
    numeric_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        errorTitle='Invalid Input',
        error='Please enter a positive number',
        promptTitle='Lessons per week',
        prompt='Enter number of lessons (minimum: 0)'
    )
    ws.add_data_validation(numeric_validation)
    numeric_validation.add('G2:H1000')
    
    # Add sample data
    sample_data = [
        "John",  # name
        "Smith",  # surname
        "john.smith@school.com",  # email
        "+1234567890",  # phone
        ", ".join(list(user_subjects)[:2]) if len(user_subjects) >= 2 else "Math, Physics",  # subjects
        ", ".join(list(user_grades)[:2]) if len(user_grades) >= 2 else "Grade 1, Grade 2",  # grades
        "20",  # min lessons
        "30",  # max lessons
        "T001"  # teacher_id
    ]
    
    # Write sample data
    for col_idx, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Add Instructions sheet
    instructions = wb.create_sheet("Instructions")
    instructions_text = [
        ["Teacher Template Instructions", ""],
        ["Required Fields:", "Fields marked with * are mandatory"],
        ["Grades (Column F):", """
• Click cell to see dropdown with pre-made grade combinations
• Select from dropdown OR type your own combination
• Multiple grades should be comma-separated
• Example: Grade 1, Grade 2
• All grades must exist in the system"""],
        ["Subjects (Column E):", """
• Click cell to see dropdown of existing subjects
• Can select from dropdown OR type new subjects
• Separate multiple subjects with commas
• Example: Math, Physics, Chemistry
• New subjects will be created automatically"""],
        ["Lessons per Week:", """
• Both minimum and maximum values are required
• Must be positive numbers
• Maximum must be greater than or equal to minimum"""],
        ["Important Notes:", """
• The dropdown for grades shows common combinations
• You can still type any valid combination of grades
• New subjects can be added by typing them
• Email format must be valid
• Teacher ID is optional"""]
    ]
    
    for row_idx, (title, content) in enumerate(instructions_text, 1):
        instructions.cell(row=row_idx, column=1, value=title).font = Font(bold=True)
        instructions.cell(row=row_idx, column=2, value=content)
        instructions.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
    
    instructions.column_dimensions['A'].width = 25
    instructions.column_dimensions['B'].width = 70
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=teacher_template.xlsx'
    
    wb.save(response)
    return response




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def process_teacher_template(request):
    """Process uploaded Excel template and create Teacher records"""
    print("hi")
    if 'file' not in request.FILES:
        return Response({"error": "No file uploaded"}, status=400)
    
    try:
        # Read Excel file
        excel_file = request.FILES['file']
        df = pd.read_excel(excel_file, sheet_name="Teachers")
        
        # Basic validation of required columns
        required_columns = ['name', 'qualified_subjects', 'grades', 
                          'min_lessons_per_week', 'max_lessons_per_week']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return Response({
                "error": f"Missing required columns: {', '.join(missing_columns)}"
            }, status=400)
        
        # Initialize response data
        response_data = {
            "success": [],
            "errors": [],
            "warnings": []
        }
        
        # Process each row
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Basic validation
                    if pd.isna(row['name']) or str(row['name']).strip() == '':
                        raise ValueError("Name is required")
                    
                    if pd.isna(row['min_lessons_per_week']) or pd.isna(row['max_lessons_per_week']):
                        raise ValueError("Minimum and maximum lessons are required")
                    
                    if row['min_lessons_per_week'] > row['max_lessons_per_week']:
                        raise ValueError("Minimum lessons cannot be greater than maximum lessons")
                    
                    # Process and validate grades
                    grade_names = [g.strip() for g in str(row['grades']).split(',') if g.strip()]
                    grades = []
                    for grade_name in grade_names:
                        grade = Grade.objects.filter(
                            name=grade_name,
                            school=request.user
                        ).first()
                        if not grade:
                            raise ValueError(f"Grade '{grade_name}' not found")
                        grades.append(grade)
                    
                    # Process subjects with fuzzy matching
                    subject_names = [s.strip() for s in str(row['qualified_subjects']).split(',') if s.strip()]
                    subjects = []
                    for subject_name in subject_names:
                        # Try exact match first
                        subject = Subject.objects.filter(
                            name=subject_name,
                            school=request.user
                        ).first()
                        
                        if not subject:
                            # Try fuzzy matching if exact match fails
                            existing_subjects = Subject.objects.filter(school=request.user)
                            best_match = None
                            best_ratio = 0
                            
                            for existing_subject in existing_subjects:
                                ratio = fuzz.ratio(subject_name.lower(), existing_subject.name.lower())
                                if ratio > 90 and ratio > best_ratio:  # 90% similarity threshold
                                    best_match = existing_subject
                                    best_ratio = ratio
                            
                            if best_match:
                                subject = best_match
                                response_data["warnings"].append(
                                    f"Row {index + 2}: Matched '{subject_name}' to existing subject '{best_match.name}'"
                                )
                            else:
                                # Create new subject if no match found
                                subject = Subject.objects.create(
                                    name=subject_name,
                                    school=request.user
                                )
                                response_data["warnings"].append(
                                    f"Row {index + 2}: Created new subject '{subject_name}'"
                                )
                        
                        subjects.append(subject)
                    
                    # Create teacher
                    teacher = Teacher.objects.create(
                        school=request.user,
                        name=str(row['name']).strip(),
                        surname=str(row['surname']).strip() if not pd.isna(row['surname']) else None,
                        email=str(row['email']).strip() if not pd.isna(row['email']) else None,
                        phone=str(row['phone']).strip() if not pd.isna(row['phone']) else None,
                        min_lessons_per_week=int(row['min_lessons_per_week']),
                        max_lessons_per_week=int(row['max_lessons_per_week']),
                        teacher_id=str(row['teacher_id']).strip() if not pd.isna(row['teacher_id']) else None
                    )
                    
                    # Add many-to-many relationships
                    teacher.qualified_subjects.set(subjects)
                    teacher.grades.set(grades)
                    
                    response_data["success"].append(f"Row {index + 2}: Created teacher {teacher.name}")
                    
                except Exception as e:
                    response_data["errors"].append(f"Row {index + 2}: {str(e)}")
                    raise  # Re-raise to trigger rollback
        
        # If we get here, all rows were processed successfully
        return Response(response_data, status=200)
                    
    except Exception as e:
        return Response({
            "error": f"Error processing template: {str(e)}",
            "details": response_data
        }, status=400)